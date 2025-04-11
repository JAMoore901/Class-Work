# File: clearClutter.py
# Student: Jessica Moore 367901
# Course: SODV 2402 FS2
# Date: April 13 2025 
# Description: This script cleans up a downloads folder by removing executable, 
# image, and zip files, and generates a report of deleted files.
#  if using the test file generator script = test_folder = downloads_path / "Downloads_To_Delete"

import os
import openpyxl
from datetime import datetime

def clean_folder(folder_path):
    """Remove specified file types and return deleted files info."""
    deleted_files = []
    
    try:
        files_to_delete = [
            f for f in os.listdir(folder_path) 
            if os.path.isfile(os.path.join(folder_path, f)) and
            f.lower().endswith(('.png', '.jpg', '.jpeg', '.exe', '.img', '.zip'))
        ]
        
        for filename in files_to_delete:
            try:
                file_path = os.path.join(folder_path, filename)
                size = os.path.getsize(file_path)
                os.remove(file_path)
                
                deleted_files.append({
                    'name': filename,
                    'type': os.path.splitext(filename)[1][1:].upper(),
                    'size': size,
                    'time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                })
                
                print(f"Removed: {filename}")
                
            except Exception as e:
                print(f"Error deleting {filename}: {str(e)}")
                
    except Exception as e:
        print(f"Folder error: {str(e)}")
    
    return deleted_files

def format_size(size):
    """Convert bytes to readable format (KB, MB, GB)."""
    for unit in ['KB', 'MB', 'GB']:
        if size < 1024.0:
            return f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} GB"

def update_excel_report(folder_path, deleted_files):
    """Update Excel report with only requested columns."""
    report_path = os.path.join(folder_path, 'deletedFiles.xlsx')
    
    try:
        if os.path.exists(report_path):
            wb = openpyxl.load_workbook(report_path)
            sheet = wb.active
            next_row = sheet.max_row + 1
        else:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Deleted Files"
            sheet.append(["Filename", "Type", "Readable Size", "Deleted On"])
            next_row = 2
        
        for file in deleted_files:
            sheet.append([
                file['name'],
                file['type'],
                format_size(file['size']),
                file['time']
            ])

        for col in sheet.columns:
            max_len = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 2
        
        wb.save(report_path)
        print(f"\nReport saved to: {report_path}")
        
    except Exception as e:
        print(f"Error saving report: {str(e)}")

def main():
    folder_path = input("Enter folder path to clean: ").strip()
    
    if not os.path.isdir(folder_path):
        print("Invalid folder path")
        return
    
    deleted_files = clean_folder(folder_path)
    
    if deleted_files:
        print(f"\nDeleted {len(deleted_files)} files")
        update_excel_report(folder_path, deleted_files)
    else:
        print("\nNo files were deleted")

if __name__ == "__main__":
    try:
        import openpyxl
        main()
    except ImportError:
        print("Please install openpyxl: pip install openpyxl")